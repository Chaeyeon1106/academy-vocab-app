"""
학원 단어시험 진도 관리 API

로컬 실행(터미널): uvicorn app.main:app --reload

백엔드: http://127.0.0.1:8000/docs 

프론트엔드: http://127.0.0.1:8000/dashboard
"""

from typing import List, Optional
import re
import io
import os
import uuid
from datetime import date, timedelta

from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import HTMLResponse, FileResponse
from sqlalchemy.orm import Session
import openpyxl

from . import models, schemas
from .database import engine, get_db

KOREAN_WEEKDAYS = {"월": 0, "화": 1, "수": 2, "목": 3, "금": 4, "토": 5, "일": 6}

FINAL_REVIEW_DIR = os.path.join(os.path.dirname(__file__), "..", "final_review_uploads")
os.makedirs(FINAL_REVIEW_DIR, exist_ok=True)


def next_attend_date(exam_date_value, attend_days_str):
    if not attend_days_str:
        return None
    days = {KOREAN_WEEKDAYS[ch] for ch in attend_days_str if ch in KOREAN_WEEKDAYS}
    if not days:
        return None
    for offset in range(1, 8):
        candidate = exam_date_value + timedelta(days=offset)
        if candidate.weekday() in days:
            return candidate
    return None

def generate_sessions(block_size, start_day=1, count=30):
    sessions = []
    current_day = start_day
    next_threshold = ((current_day - 1) // 10 + 1) * 10

    while len(sessions) < count:
        block_end = current_day + block_size - 1
        sessions.append(("regular", current_day, block_end))
        current_day = block_end + 1

        if block_end >= next_threshold:
            sessions.append(("final_review", next_threshold - 9, next_threshold))
            next_threshold += 10

    return sessions


def session_index_for_date(start_date, attend_days_str, on_date, absent_dates=frozenset()):
    if not attend_days_str or on_date < start_date:
        return None
    days = {KOREAN_WEEKDAYS[ch] for ch in attend_days_str if ch in KOREAN_WEEKDAYS}
    if on_date.weekday() not in days:
        return None
    if on_date in absent_dates:
        return None

    count = 0
    d = start_date
    while d <= on_date:
        if d.weekday() in days and d not in absent_dates:
            if d == on_date:
                return count
            count += 1
        d += timedelta(days=1)
    return None


def compute_today_session(student, on_date, total_days=None, absent_dates=frozenset()):
    if not (student.plan_days_per_test and student.plan_start_date and student.attend_days):
        return None
    start_day = student.plan_start_day or 1
    start_round = student.plan_start_round or 1
    index = session_index_for_date(student.plan_start_date, student.attend_days, on_date, absent_dates)
    if index is None:
        return None
    sessions = generate_sessions(student.plan_days_per_test, start_day, count=index + 1)
    session_type, s_day, e_day = sessions[index]

    round_number = start_round
    display_start = s_day
    display_end = e_day
    if total_days:
        round_offset, day_in_round = divmod(s_day - 1, total_days)
        round_number = start_round + round_offset
        display_start = day_in_round + 1
        display_end = min(display_start + (e_day - s_day), total_days)

    return {"type": session_type, "start_day": display_start, "end_day": display_end, "round": round_number}


RETEST_SHEET_STYLE = """
    body { font-family: sans-serif; padding: 24px; }
    h1 { font-size: 20px; margin-bottom: 4px; }
    h2 { font-size: 15px; margin: 0 0 4px; }
    .meta { margin-bottom: 16px; color: #333; }
    .book { font-weight: normal; color: #666; font-size: 13px; }
    .class { font-weight: normal; color: #666; font-size: 14px; }
    .day-pair { display: flex; gap: 24px; margin-top: 20px; align-items: flex-start; }
    .day-col { flex: 1; min-width: 0; }
    table { border-collapse: collapse; width: 100%; }
    th, td { border: 1px solid #999; padding: 4px 8px; text-align: left; font-size: 14px; }
    th { background: #f0f0f0; }
    .no { width: 34px; text-align: center; }
    .check { width: 50px; }
    .page { page-break-after: always; padding-bottom: 24px; }
    @media print { button { display: none; } }
"""


def render_word_table(words):
    rows_html = "".join(
        f"<tr><td class='no'>{i+1}</td><td class='word'>{w.word}</td><td class='check'></td></tr>"
        for i, w in enumerate(words)
    )
    return f"""<table>
        <tr><th class="no">No.</th><th>단어</th><th class="check">통과</th></tr>
        {rows_html}
    </table>"""


def render_day_sections(sections):
    """sections: list of (title_html, words) tuples. Renders two sections per row, side by side."""
    html = ""
    for i in range(0, len(sections), 2):
        cols_html = ""
        for title, words in sections[i:i + 2]:
            cols_html += f"""
            <div class="day-col">
                <h2>{title}</h2>
                {render_word_table(words)}
            </div>
            """
        html += f'<div class="day-pair">{cols_html}</div>'
    return html


def get_ordered_words(db, book_id, start_day, end_day, round_number=1):
    return (
        db.query(models.Word)
        .filter(
            models.Word.book_id == book_id,
            models.Word.day_number >= start_day,
            models.Word.day_number <= end_day,
            models.Word.round_number == round_number,
        )
        .order_by(models.Word.day_number, models.Word.id)
        .all()
    )

# 서버 처음 켤 때 테이블이 없으면 자동으로 생성해줍니다.
models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="학원 단어시험 진도 관리 API")


# ---------------------------------------------------------------
# Dashboard (HTML)
# ---------------------------------------------------------------
@app.get("/dashboard", response_class=HTMLResponse)
def dashboard():
    html = """
    <html>
    <head>
    <meta charset="utf-8">
    <title>학원 단어시험 관리</title>
    <style>
        * { box-sizing: border-box; }
        body { font-family: -apple-system, "Segoe UI", sans-serif; padding: 24px; max-width: 1000px;
               margin: 0 auto; background: #f8fafc; color: #1f2937; }
        h1 { font-size: 22px; color: #111827; }
        h2 { font-size: 16px; margin-top: 28px; margin-bottom: 8px; color: #1f2937;
             border-left: 4px solid #2563eb; padding-left: 8px; }
        p { font-size: 13px; }
        table { border-collapse: collapse; width: 100%; margin-top: 8px; background: #fff; }
        th, td { border: 1px solid #e5e7eb; padding: 8px 10px; text-align: left; font-size: 14px; }
        th { background: #f3f4f6; }
        a.btn { display: inline-block; background: #2563eb; color: white; padding: 4px 10px;
                border-radius: 4px; text-decoration: none; font-size: 13px; }
        a.btn:hover { background: #1d4ed8; }
        .empty { color: #9ca3af; padding: 12px 0; }
        button.refresh { padding: 6px 14px; cursor: pointer; border: 1px solid #d1d5db;
                          border-radius: 6px; background: #fff; }
        input, select, textarea { border: 1px solid #d1d5db; border-radius: 6px; padding: 5px 8px; font-size: 13px; }
        form { background: #fff; border: 1px solid #e5e7eb; border-radius: 8px; padding: 14px; margin-top: 8px; }
        form button { border: 1px solid #2563eb; background: #2563eb; color: #fff; border-radius: 6px; }
        form button:hover { background: #1d4ed8; }

        .tabs { display: flex; gap: 4px; margin-top: 18px; border-bottom: 2px solid #e5e7eb; }
        .tab-btn { padding: 10px 20px; border: none; background: none; cursor: pointer;
                   font-size: 14px; color: #6b7280; border-bottom: 2px solid transparent; margin-bottom: -2px; }
        .tab-btn.active { color: #2563eb; border-bottom-color: #2563eb; font-weight: 600; }
        .tab-panel { display: none; padding-top: 16px; }
        .tab-panel.active { display: block; }

        .calendar-grid { display: grid; grid-template-columns: repeat(7, 1fr); gap: 4px; margin-top: 8px; }
        .cal-weekday { text-align: center; font-size: 12px; color: #9ca3af; padding-bottom: 4px; }
        .cal-cell { border: 1px solid #e5e7eb; border-radius: 6px; min-height: 60px; padding: 6px; background: #fff; }
        .cal-cell.empty { border: none; background: none; }
        .cal-cell.has-data { cursor: pointer; background: #eff6ff; border-color: #bfdbfe; }
        .cal-cell.has-data:hover { background: #dbeafe; }
        .cal-cell.selected { outline: 2px solid #2563eb; }
        .cal-cell .cal-date { font-size: 12px; color: #6b7280; }
        .cal-cell .cal-count { font-size: 11px; color: #2563eb; margin-top: 4px; font-weight: 600; }
    </style>
    </head>
    <body>
        <h1>학원 단어시험 관리</h1>
        <div class="tabs">
            <button class="tab-btn active" data-tab="today">오늘 · 프린트</button>
            <button class="tab-btn" data-tab="manage">관리</button>
            <button class="tab-btn" data-tab="history">과거 기록</button>
        </div>

        <div id="tab-manage" class="tab-panel">
        <h2>교사 추가</h2>
        <form id="teacher-add-form" style="display:flex; gap:8px; flex-wrap: wrap; align-items:end; margin-top:8px;">
            <label style="font-size:13px;">이름<br>
                <input type="text" id="t-name" style="width:100px;" required>
            </label>
            <label style="font-size:13px;">아이디<br>
                <input type="text" id="t-username" style="width:100px;" required>
            </label>
            <label style="font-size:13px;">비밀번호<br>
                <input type="password" id="t-password" style="width:100px;" required>
            </label>
            <button type="button" id="t-submit" style="padding:6px 14px; cursor:pointer;">등록</button>
            <span id="t-status" style="font-size:13px; color:#2563eb;"></span>
        </form>

        <h2>단어책 추가</h2>
        <form id="book-add-form" style="display:flex; gap:8px; flex-wrap: wrap; align-items:end; margin-top:8px;">
            <label style="font-size:13px;">출판사 (선택)<br>
                <input type="text" id="bk-publisher" style="width:100px;">
            </label>
            <label style="font-size:13px;">책 이름 (한 줄에 하나씩)<br>
                <textarea id="bk-names" rows="4" style="width:220px;" required></textarea>
            </label>
            <button type="button" id="bk-submit" style="padding:6px 14px; cursor:pointer;">추가</button>
            <span id="bk-status" style="font-size:13px; color:#2563eb;"></span>
        </form>

        <h2>단어 엑셀 업로드</h2>
        <p style="color:#888; font-size:13px; margin: 4px 0 8px;">단어책마다 엑셀 파일을 업로드해야 단어/Day 범위/통과 기준이 등록되고, 재시험지와 자동판정이 동작해요.</p>
        <form id="excel-upload-form" style="display:flex; gap:8px; flex-wrap: wrap; align-items:end; margin-top:8px;">
            <label style="font-size:13px;">단어책<br>
                <select id="ex-book" style="width:150px;" required>
                    <option value="">선택</option>
                </select>
            </label>
            <label style="font-size:13px;">시트 이름<br>
                <select id="ex-sheet" style="width:110px;">
                    <option value="학원용" selected>학원용</option>
                    <option value="">기본(첫 시트)</option>
                    <option value="가정용">가정용</option>
                    <option value="원본">원본</option>
                </select>
            </label>
            <label style="font-size:13px;">엑셀 파일<br>
                <input type="file" id="ex-file" accept=".xlsx,.xls" required>
            </label>
            <button type="button" id="ex-submit" style="padding:6px 14px; cursor:pointer;">업로드</button>
            <span id="ex-status" style="font-size:13px; color:#2563eb;"></span>
        </form>

        <h2>최종시험지 업로드</h2>
        <p style="color:#888; font-size:13px; margin: 4px 0 8px;">hwp/pdf 등 이미 만들어진 최종복습 시험지 파일을 그대로 올려두면, 오늘 시험 범위에서 바로 열어 인쇄할 수 있어요.</p>
        <form id="frf-upload-form" style="display:flex; gap:8px; flex-wrap: wrap; align-items:end; margin-top:8px;">
            <label style="font-size:13px;">단어책<br>
                <select id="frf-book" style="width:150px;" required>
                    <option value="">선택</option>
                </select>
            </label>
            <label style="font-size:13px;">시작 Day<br>
                <input type="number" id="frf-start-day" min="1" style="width:70px;" required>
            </label>
            <label style="font-size:13px;">끝 Day<br>
                <input type="number" id="frf-end-day" min="1" style="width:70px;" required>
            </label>
            <label style="font-size:13px;">버전<br>
                <input type="text" id="frf-version" value="A" style="width:50px;">
            </label>
            <label style="font-size:13px;">파일 (hwp/pdf 등)<br>
                <input type="file" id="frf-file" required>
            </label>
            <button type="button" id="frf-submit" style="padding:6px 14px; cursor:pointer;">업로드</button>
            <span id="frf-status" style="font-size:13px; color:#2563eb;"></span>
        </form>

        <h2>학생 추가</h2>
        <form id="student-add-form" style="display:flex; gap:8px; flex-wrap: wrap; align-items:end; margin-top:8px;">
            <label style="font-size:13px;">이름<br>
                <input type="text" id="s-name" style="width:100px;" required>
            </label>
            <label style="font-size:13px;">반<br>
                <input type="text" id="s-class" style="width:100px;">
            </label>
            <label style="font-size:13px;">요일<br>
                <span id="s-weekdays" style="display:flex; gap:4px; padding-top:4px;">
                    <label><input type="checkbox" value="월">월</label>
                    <label><input type="checkbox" value="화">화</label>
                    <label><input type="checkbox" value="수">수</label>
                    <label><input type="checkbox" value="목">목</label>
                    <label><input type="checkbox" value="금">금</label>
                    <label><input type="checkbox" value="토">토</label>
                    <label><input type="checkbox" value="일">일</label>
                </span>
            </label>
            <label style="font-size:13px;">담임<br>
                <select id="s-teacher" style="width:100px;">
                    <option value="">선택 안함</option>
                </select>
            </label>
            <label style="font-size:13px;">단어책<br>
                <select id="s-book" style="width:120px;">
                    <option value="">선택 안함</option>
                </select>
            </label>
            <button type="button" id="s-submit" style="padding:6px 14px; cursor:pointer;">추가</button>
            <span id="s-status" style="font-size:13px; color:#2563eb;"></span>
        </form>

        <h2>자동 스케줄 설정</h2>
        <p style="color:#888; font-size:13px; margin: 4px 0 8px;">한 번만 설정하면, 그 다음부터 등원일마다 시험 범위가 자동 계산돼요.</p>
        <form id="plan-form" style="display:flex; gap:8px; flex-wrap: wrap; align-items:end;">
            <label style="font-size:13px;">학생<br>
                <select id="p-student" required></select>
            </label>
            <label style="font-size:13px;">며칠씩<br>
                <input type="number" id="p-days" min="1" value="5" style="width:60px;" required>
            </label>
            <label style="font-size:13px;">시작 Day<br>
                <input type="number" id="p-start-day" min="1" value="1" style="width:70px;" required>
            </label>
            <label style="font-size:13px;">시작 회독<br>
                <input type="number" id="p-start-round" min="1" value="1" style="width:60px;" required>
            </label>
            <label style="font-size:13px;">시작 날짜<br>
                <input type="date" id="p-start-date" required>
            </label>
            <button type="submit" style="padding:6px 14px; cursor:pointer;">설정</button>
            <span id="p-status" style="font-size:13px; color:#2563eb;"></span>
        </form>
        </div>

        <div id="tab-today" class="tab-panel active">
        <div style="display:flex; align-items:center; gap:12px; margin-top:8px; flex-wrap: wrap;">
            <button class="refresh" onclick="loadData()">새로고침</button>
            <label style="font-size:14px;">반 선택:
                <select id="class-filter" onchange="renderAll(); loadTodaySessions();">
                    <option value="">전체</option>
                </select>
            </label>
            <label style="font-size:14px;">
                <input type="checkbox" id="due-today-only" checked onchange="renderAll()">
                오늘 인쇄 대상만 보기
            </label>
            <a class="btn" id="class-print-link" href="#" target="_blank" style="display:none;">이 반 오늘 재시험지 일괄 인쇄</a>
        </div>

        <h2>시험 점수 입력</h2>
        <form id="exam-form" style="display:flex; gap:8px; flex-wrap: wrap; align-items:end; margin-top:8px;">
            <label style="font-size:13px;">학생<br>
                <select id="f-student" required></select>
            </label>
            <label style="font-size:13px;">회독<br>
                <input type="number" id="f-round" value="1" min="1" style="width:60px;" required>
            </label>
            <label style="font-size:13px;">시작 Day<br>
                <input type="number" id="f-start-day" min="1" style="width:70px;" required>
            </label>
            <label style="font-size:13px;">끝 Day<br>
                <input type="number" id="f-end-day" min="1" style="width:70px;" required>
            </label>
            <label style="font-size:13px;">시험일<br>
                <input type="date" id="f-date" required>
            </label>
            <button type="submit" style="padding:6px 14px; cursor:pointer;">전체 등록</button>
            <span id="f-status" style="font-size:13px; color:#2563eb;"></span>
            <div style="width:100%; font-size:12px; color:#888;" id="f-hint"></div>
            <div style="width:100%;" id="f-day-rows"></div>
        </form>

        <h2>오늘 시험 범위 (자동계산)</h2>
        <div id="today-empty" class="empty" style="display:none;">반을 선택하면 오늘 볼 시험 범위가 나와요.</div>
        <table id="today-table" style="display:none;">
            <tr><th>학생</th><th>출석</th><th>단어책</th><th>회독</th><th>종류</th><th>범위</th><th>최종시험지</th></tr>
            <tbody id="today-body"></tbody>
        </table>

        <h2>재시험 대상자</h2>
        <div id="retest-empty" class="empty" style="display:none;">재시험 대상자가 없습니다.</div>
        <table id="retest-table">
            <tr><th>학생</th><th>반</th><th>요일</th><th>단어책</th><th>회독</th><th>범위</th><th>점수</th><th>인쇄 예정일</th><th>재시험지</th></tr>
            <tbody id="retest-body"></tbody>
        </table>

        <h2>전체 학생 진도</h2>
        <table>
            <tr><th>이름</th><th>반</th><th>요일</th><th>담임</th><th>현재 단어책</th><th>재시험지(통합)</th></tr>
            <tbody id="students-body"></tbody>
        </table>
        </div>

        <div id="tab-history" class="tab-panel">
        <h2>등록된 선생님</h2>
        <table>
            <tr><th>이름</th><th>아이디</th></tr>
            <tbody id="roster-teachers-body"></tbody>
        </table>

        <h2>등록된 학생</h2>
        <table>
            <tr><th>이름</th><th>반</th><th>요일</th><th>담임</th><th>단어책 변경</th><th>삭제</th></tr>
            <tbody id="roster-students-body"></tbody>
        </table>

        <h2>날짜별 시험 기록</h2>
        <div style="display:flex; align-items:center; gap:12px; margin-bottom:4px;">
            <button id="cal-prev" style="padding:4px 10px; cursor:pointer; border:1px solid #d1d5db; border-radius:6px; background:#fff;">‹ 이전달</button>
            <span id="cal-label" style="font-weight:600; font-size:16px;"></span>
            <button id="cal-next" style="padding:4px 10px; cursor:pointer; border:1px solid #d1d5db; border-radius:6px; background:#fff;">다음달 ›</button>
        </div>
        <div id="calendar-grid" class="calendar-grid"></div>
        <div id="day-detail" style="margin-top:20px;"></div>
        </div>

    <script>
        let allData = { teachers: [], books: [], students: [], candidates: [], examRecords: [], finalReviewFiles: [] };

        function todayStr() {
            const d = new Date();
            return d.toISOString().slice(0, 10);
        }

        async function loadData() {
            document.getElementById('f-date').value = todayStr();

            const [teachers, books, students] = await Promise.all([
                fetch('/teachers').then(r => r.json()),
                fetch('/books').then(r => r.json()),
                fetch('/students').then(r => r.json()),
            ]);
            const candidatesAll = await fetch('/exam-records/retest-candidates').then(r => r.json());
            const examRecords = await fetch('/exam-records').then(r => r.json());
            const finalReviewFiles = await fetch('/final-review-files').then(r => r.json());
            allData = { teachers, books, students, candidates: candidatesAll, examRecords, finalReviewFiles };

            const classSelect = document.getElementById('class-filter');
            const currentValue = classSelect.value;
            const classNames = [...new Set(students.map(s => s.class_name).filter(Boolean))].sort();
            classSelect.innerHTML = '<option value="">전체</option>' +
                classNames.map(c => `<option value="${c}">${c}</option>`).join('');
            classSelect.value = classNames.includes(currentValue) ? currentValue : '';

            const studentSelect = document.getElementById('f-student');
            const studentSelectValue = studentSelect.value;
            studentSelect.innerHTML = students.map(s => `<option value="${s.id}">${s.name} (${s.class_name || '반없음'})</option>`).join('');
            if (students.some(s => String(s.id) === studentSelectValue)) {
                studentSelect.value = studentSelectValue;
            }

            const planStudentSelect = document.getElementById('p-student');
            planStudentSelect.innerHTML = students.map(s => `<option value="${s.id}">${s.name} (${s.class_name || '반없음'})</option>`).join('');
            document.getElementById('p-start-date').value = todayStr();

            const sTeacherSelect = document.getElementById('s-teacher');
            const sTeacherValue = sTeacherSelect.value;
            sTeacherSelect.innerHTML = '<option value="">선택 안함</option>' +
                teachers.map(t => `<option value="${t.id}">${t.name}</option>`).join('');
            sTeacherSelect.value = sTeacherValue;

            const sBookSelect = document.getElementById('s-book');
            const sBookValue = sBookSelect.value;
            sBookSelect.innerHTML = '<option value="">선택 안함</option>' +
                books.map(b => `<option value="${b.id}">${b.name}</option>`).join('');
            sBookSelect.value = sBookValue;

            const exBookSelect = document.getElementById('ex-book');
            const exBookValue = exBookSelect.value;
            exBookSelect.innerHTML = '<option value="">선택</option>' +
                books.map(b => `<option value="${b.id}">${b.name}</option>`).join('');
            exBookSelect.value = exBookValue;

            const frfBookSelect = document.getElementById('frf-book');
            const frfBookValue = frfBookSelect.value;
            frfBookSelect.innerHTML = '<option value="">선택</option>' +
                books.map(b => `<option value="${b.id}">${b.name}</option>`).join('');
            frfBookSelect.value = frfBookValue;

            renderAll();
            loadTodaySessions();
            renderRoster();
            if (students.length) {
                examAutoFillDay();
            }
        }

        function renderRoster() {
            const { teachers, students, books } = allData;
            const bookMap = Object.fromEntries(books.map(b => [b.id, b.name]));
            const teacherMap = Object.fromEntries(teachers.map(t => [t.id, t.name]));

            document.getElementById('roster-teachers-body').innerHTML = teachers.map(t => `
                <tr><td>${t.name}</td><td>${t.username}</td></tr>
            `).join('');

            document.getElementById('roster-students-body').innerHTML = students.map(s => `
                <tr>
                    <td>${s.name}</td>
                    <td>${s.class_name || '-'}</td>
                    <td>${s.attend_days || '-'}</td>
                    <td>${teacherMap[s.teacher_id] || '-'}</td>
                    <td>
                        <select class="roster-book-select" data-student-id="${s.id}" style="width:130px;">
                            <option value="">단어책 없음</option>
                            ${books.map(b => `<option value="${b.id}" ${b.id === s.current_book_id ? 'selected' : ''}>${b.name}</option>`).join('')}
                        </select>
                    </td>
                    <td><button type="button" class="roster-delete-btn" data-student-id="${s.id}" data-student-name="${s.name}" style="padding:4px 10px; cursor:pointer;">삭제</button></td>
                </tr>
            `).join('');

            document.querySelectorAll('.roster-book-select').forEach(sel => {
                sel.addEventListener('change', async () => {
                    const studentId = sel.dataset.studentId;
                    const bookId = sel.value ? parseInt(sel.value) : null;
                    try {
                        const res = await fetch(`/students/${studentId}`, {
                            method: 'PATCH',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({ current_book_id: bookId }),
                        });
                        if (res.ok) {
                            loadData();
                        } else {
                            alert('단어책 변경 실패');
                        }
                    } catch (err) {
                        alert('서버에 연결할 수 없습니다.');
                        console.error(err);
                    }
                });
            });

            document.querySelectorAll('.roster-delete-btn').forEach(btn => {
                btn.addEventListener('click', async () => {
                    const studentId = btn.dataset.studentId;
                    const studentName = btn.dataset.studentName;
                    if (!confirm(`"${studentName}" 학생을 삭제할까요? 이 학생의 시험 기록도 함께 삭제될 수 있습니다.`)) return;
                    try {
                        const res = await fetch(`/students/${studentId}`, { method: 'DELETE' });
                        if (res.ok) {
                            loadData();
                        } else {
                            alert('삭제 실패');
                        }
                    } catch (err) {
                        alert('서버에 연결할 수 없습니다.');
                        console.error(err);
                    }
                });
            });
        }

        async function loadTodaySessions() {
            const selectedClass = document.getElementById('class-filter').value;
            const todayTable = document.getElementById('today-table');
            const todayEmpty = document.getElementById('today-empty');
            const todayBody = document.getElementById('today-body');

            if (!selectedClass) {
                todayTable.style.display = 'none';
                todayEmpty.style.display = 'block';
                todayEmpty.textContent = '반을 선택하면 오늘 볼 시험 범위가 나와요.';
                return;
            }

            const [res, attRes] = await Promise.all([
                fetch(`/classes/${encodeURIComponent(selectedClass)}/today-sessions`),
                fetch(`/classes/${encodeURIComponent(selectedClass)}/attendance`),
            ]);
            const data = await res.json();
            const attData = await attRes.json();
            const sessionByStudent = Object.fromEntries(data.sessions.map(s => [s.student_id, s]));

            const classStudents = allData.students.filter(s => s.class_name === selectedClass);
            if (!classStudents.length) {
                todayTable.style.display = 'none';
                todayEmpty.style.display = 'block';
                todayEmpty.textContent = '이 반에 등록된 학생이 없습니다.';
                return;
            }

            const bookMap = Object.fromEntries(allData.books.map(b => [b.id, b.name]));

            todayEmpty.style.display = 'none';
            todayTable.style.display = 'table';
            todayBody.innerHTML = classStudents.map(student => {
                const s = sessionByStudent[student.id];
                const present = attData.attendance[student.id] !== false;
                let finalFileCell = '-';
                if (s && s.type === 'final_review') {
                    const match = allData.finalReviewFiles.find(f =>
                        f.book_id === student.current_book_id && f.start_day === s.start_day && f.end_day === s.end_day
                    );
                    finalFileCell = match
                        ? `<a class="btn" href="/final-review-files/${match.id}/download" target="_blank">최종시험지 열기</a>`
                        : '<span style="color:#f59e0b;">파일 없음</span>';
                }
                return `
                <tr>
                    <td>${student.name}</td>
                    <td><input type="checkbox" class="att-checkbox" data-student-id="${student.id}" ${present ? 'checked' : ''}></td>
                    <td>${bookMap[student.current_book_id] || '-'}</td>
                    <td>${s ? `${s.round}회독` : '-'}</td>
                    <td>${s ? (s.type === 'final_review' ? '최종복습' : '정규') : '-'}</td>
                    <td>${s ? `Day ${s.start_day}~${s.end_day}` : '오늘 해당 없음'}</td>
                    <td>${finalFileCell}</td>
                </tr>
                `;
            }).join('');

            document.querySelectorAll('.att-checkbox').forEach(cb => {
                cb.addEventListener('change', async () => {
                    const studentId = parseInt(cb.dataset.studentId);
                    try {
                        const res = await fetch('/attendance', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({ student_id: studentId, date: todayStr(), present: cb.checked }),
                        });
                        if (res.ok) {
                            loadTodaySessions();
                        } else {
                            alert('출석 체크 실패');
                            cb.checked = !cb.checked;
                        }
                    } catch (err) {
                        alert('서버에 연결할 수 없습니다.');
                        cb.checked = !cb.checked;
                        console.error(err);
                    }
                });
            });
        }

        document.getElementById('t-submit').addEventListener('click', async () => {
            const body = {
                name: document.getElementById('t-name').value.trim(),
                username: document.getElementById('t-username').value.trim(),
                password: document.getElementById('t-password').value,
            };
            if (!body.name || !body.username || !body.password) {
                document.getElementById('t-status').textContent = '이름/아이디/비밀번호를 모두 입력해주세요.';
                return;
            }
            try {
                const res = await fetch('/teachers', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify(body),
                });
                if (res.ok) {
                    document.getElementById('t-status').textContent = '등록 완료!';
                    document.getElementById('t-name').value = '';
                    document.getElementById('t-username').value = '';
                    document.getElementById('t-password').value = '';
                    loadData();
                } else {
                    const errText = await res.text();
                    let detail = errText;
                    try { detail = JSON.parse(errText).detail || errText; } catch (e) {}
                    document.getElementById('t-status').textContent = '등록 실패: ' + detail;
                    console.error(errText);
                }
            } catch (err) {
                document.getElementById('t-status').textContent = '서버에 연결할 수 없습니다.';
                console.error(err);
            }
        });

        document.getElementById('bk-submit').addEventListener('click', async () => {
            const names = document.getElementById('bk-names').value
                .split('\\n').map(s => s.trim()).filter(Boolean);
            const publisher = document.getElementById('bk-publisher').value.trim();
            if (!names.length) {
                document.getElementById('bk-status').textContent = '책 이름을 한 줄에 하나씩 입력해주세요.';
                return;
            }

            const body = { names: names, publisher: publisher || null };
            try {
                const res = await fetch('/books/bulk', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify(body),
                });
                if (res.ok) {
                    const added = await res.json();
                    document.getElementById('bk-status').textContent =
                        `${added.length}권 추가 완료! (${names.length - added.length}권은 중복이라 건너뜀)`;
                    document.getElementById('bk-names').value = '';
                    document.getElementById('bk-publisher').value = '';
                    loadData();
                } else {
                    const errText = await res.text();
                    let detail = errText;
                    try { detail = JSON.parse(errText).detail || errText; } catch (e) {}
                    document.getElementById('bk-status').textContent = '추가 실패: ' + detail;
                    console.error(errText);
                }
            } catch (err) {
                document.getElementById('bk-status').textContent = '서버에 연결할 수 없습니다.';
                console.error(err);
            }
        });

        document.getElementById('ex-submit').addEventListener('click', async () => {
            const bookId = document.getElementById('ex-book').value;
            const sheet = document.getElementById('ex-sheet').value.trim();
            const fileInput = document.getElementById('ex-file');
            if (!bookId || !fileInput.files.length) {
                document.getElementById('ex-status').textContent = '단어책과 파일을 선택해주세요.';
                return;
            }
            const formData = new FormData();
            formData.append('book_id', bookId);
            if (sheet) formData.append('sheet_name', sheet);
            formData.append('file', fileInput.files[0]);
            try {
                const res = await fetch('/words/upload-excel', { method: 'POST', body: formData });
                if (res.ok) {
                    const result = await res.json();
                    let msg = `업로드 완료! 새 단어 ${result.newly_inserted}개 추가 (중복 ${result.skipped_duplicates}개 건너뜀)`;
                    if (result.pass_cutoff_detected !== null && result.pass_cutoff_detected !== undefined) {
                        msg += `, 통과 기준: ${result.pass_cutoff_detected}개 이하 오답`;
                    }
                    document.getElementById('ex-status').textContent = msg;
                    fileInput.value = '';
                    document.getElementById('ex-sheet').value = '';
                    loadData();
                } else {
                    const errText = await res.text();
                    let detail = errText;
                    try { detail = JSON.parse(errText).detail || errText; } catch (e) {}
                    document.getElementById('ex-status').textContent = '업로드 실패: ' + detail;
                    console.error(errText);
                }
            } catch (err) {
                document.getElementById('ex-status').textContent = '서버에 연결할 수 없습니다.';
                console.error(err);
            }
        });

        document.getElementById('frf-submit').addEventListener('click', async () => {
            const bookId = document.getElementById('frf-book').value;
            const startDay = document.getElementById('frf-start-day').value;
            const endDay = document.getElementById('frf-end-day').value;
            const version = document.getElementById('frf-version').value.trim() || 'A';
            const fileInput = document.getElementById('frf-file');
            if (!bookId || !startDay || !endDay || !fileInput.files.length) {
                document.getElementById('frf-status').textContent = '단어책, Day 범위, 파일을 모두 입력해주세요.';
                return;
            }
            const formData = new FormData();
            formData.append('book_id', bookId);
            formData.append('start_day', startDay);
            formData.append('end_day', endDay);
            formData.append('version_label', version);
            formData.append('file', fileInput.files[0]);
            try {
                const res = await fetch('/final-review-files', { method: 'POST', body: formData });
                if (res.ok) {
                    document.getElementById('frf-status').textContent = '업로드 완료!';
                    document.getElementById('frf-start-day').value = '';
                    document.getElementById('frf-end-day').value = '';
                    document.getElementById('frf-version').value = 'A';
                    fileInput.value = '';
                    loadData();
                } else {
                    const errText = await res.text();
                    let detail = errText;
                    try { detail = JSON.parse(errText).detail || errText; } catch (e) {}
                    document.getElementById('frf-status').textContent = '업로드 실패: ' + detail;
                    console.error(errText);
                }
            } catch (err) {
                document.getElementById('frf-status').textContent = '서버에 연결할 수 없습니다.';
                console.error(err);
            }
        });

        document.getElementById('s-submit').addEventListener('click', async () => {
            const name = document.getElementById('s-name').value.trim();
            const className = document.getElementById('s-class').value.trim();
            const attendDays = [...document.querySelectorAll('#s-weekdays input:checked')]
                .map(cb => cb.value).join('');
            const teacherId = document.getElementById('s-teacher').value;
            const bookId = document.getElementById('s-book').value;

            if (!name) {
                document.getElementById('s-status').textContent = '이름을 입력해주세요.';
                return;
            }

            const body = {
                name: name,
                class_name: className || null,
                attend_days: attendDays || null,
                teacher_id: teacherId ? parseInt(teacherId) : null,
                current_book_id: bookId ? parseInt(bookId) : null,
            };
            try {
                const res = await fetch('/students', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify(body),
                });
                if (res.ok) {
                    document.getElementById('s-status').textContent = '추가 완료!';
                    document.getElementById('s-name').value = '';
                    document.getElementById('s-class').value = '';
                    document.querySelectorAll('#s-weekdays input:checked').forEach(cb => cb.checked = false);
                    document.getElementById('s-teacher').value = '';
                    document.getElementById('s-book').value = '';
                    loadData();
                } else {
                    const errText = await res.text();
                    let detail = errText;
                    try { detail = JSON.parse(errText).detail || errText; } catch (e) {}
                    document.getElementById('s-status').textContent = '추가 실패: ' + detail;
                    console.error(errText);
                }
            } catch (err) {
                document.getElementById('s-status').textContent = '서버에 연결할 수 없습니다.';
                console.error(err);
            }
        });

        document.getElementById('plan-form').addEventListener('submit', async (e) => {
            e.preventDefault();
            const studentId = parseInt(document.getElementById('p-student').value);
            const body = {
                plan_days_per_test: parseInt(document.getElementById('p-days').value),
                plan_start_day: parseInt(document.getElementById('p-start-day').value),
                plan_start_round: parseInt(document.getElementById('p-start-round').value),
                plan_start_date: document.getElementById('p-start-date').value,
            };
            const res = await fetch(`/students/${studentId}`, {
                method: 'PATCH',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify(body),
            });
            if (res.ok) {
                document.getElementById('p-status').textContent = '설정 완료!';
                loadTodaySessions();
            } else {
                document.getElementById('p-status').textContent = '설정 실패 (콘솔 확인)';
                console.error(await res.text());
            }
        });

        function renderAll() {
            const { teachers, books, students, candidates } = allData;
            const selectedClass = document.getElementById('class-filter').value;
            const dueTodayOnly = document.getElementById('due-today-only').checked;

            const teacherMap = Object.fromEntries(teachers.map(t => [t.id, t.name]));
            const bookMap = Object.fromEntries(books.map(b => [b.id, b.name]));
            const studentMap = Object.fromEntries(students.map(s => [s.id, s]));

            const filteredStudents = selectedClass
                ? students.filter(s => s.class_name === selectedClass)
                : students;
            const filteredStudentIds = new Set(filteredStudents.map(s => s.id));

            let filteredCandidates = candidates.filter(c => filteredStudentIds.has(c.student_id));
            if (dueTodayOnly) {
                filteredCandidates = filteredCandidates.filter(c => c.due_date === todayStr());
            }

            const printLink = document.getElementById('class-print-link');
            if (selectedClass) {
                printLink.href = `/classes/${encodeURIComponent(selectedClass)}/retest-print-queue`;
                printLink.style.display = 'inline-block';
            } else {
                printLink.style.display = 'none';
            }

            const retestBody = document.getElementById('retest-body');
            retestBody.innerHTML = '';
            document.getElementById('retest-empty').style.display = filteredCandidates.length ? 'none' : 'block';
            filteredCandidates.forEach(c => {
                const student = studentMap[c.student_id];
                const tr = document.createElement('tr');
                tr.innerHTML = `
                    <td>${student ? student.name : '-'}</td>
                    <td>${student?.class_name || '-'}</td>
                    <td>${student?.attend_days || '-'}</td>
                    <td>${bookMap[c.book_id] || '-'}</td>
                    <td>${c.round_number}회독</td>
                    <td>Day ${c.start_day}~${c.end_day}</td>
                    <td>${c.score ?? '-'}/${c.total_questions ?? '-'}</td>
                    <td>${c.due_date || '-'}</td>
                    <td><a class="btn" href="/exam-records/${c.id}/retest-sheet" target="_blank">재시험지 열기</a></td>
                `;
                retestBody.appendChild(tr);
            });

            const studentsBody = document.getElementById('students-body');
            studentsBody.innerHTML = '';
            filteredStudents.forEach(s => {
                const hasRetest = candidates.some(c => c.student_id === s.id);
                const tr = document.createElement('tr');
                tr.innerHTML = `
                    <td>${s.name}</td>
                    <td>${s.class_name || '-'}</td>
                    <td>${s.attend_days || '-'}</td>
                    <td>${teacherMap[s.teacher_id] || '-'}</td>
                    <td>${bookMap[s.current_book_id] || '-'}</td>
                    <td>${hasRetest ? `<a class="btn" href="/students/${s.id}/retest-sheet" target="_blank">통합 재시험지</a>` : '-'}</td>
                `;
                studentsBody.appendChild(tr);
            });
        }

        async function examAutoFillDay() {
            const studentId = parseInt(document.getElementById('f-student').value);
            const hint = document.getElementById('f-hint');
            if (!studentId) return;
            try {
                const res = await fetch(`/students/${studentId}/today-session`);
                const data = await res.json();
                if (data.session) {
                    document.getElementById('f-start-day').value = data.session.start_day;
                    document.getElementById('f-end-day').value = data.session.end_day;
                    document.getElementById('f-round').value = data.session.round;
                    hint.textContent = '오늘 자동 스케줄 범위를 채웠어요. 필요하면 직접 수정하세요.';
                } else {
                    hint.textContent = '이 학생은 오늘 자동 스케줄이 없어요. Day 범위를 직접 입력해주세요.';
                }
            } catch (err) {
                console.error(err);
            }
            examBuildDayRows();
        }

        async function examBuildDayRows() {
            const studentId = parseInt(document.getElementById('f-student').value);
            const student = allData.students.find(s => s.id === studentId);
            const startDay = parseInt(document.getElementById('f-start-day').value);
            const endDay = parseInt(document.getElementById('f-end-day').value);
            const rowsContainer = document.getElementById('f-day-rows');
            const hint = document.getElementById('f-hint');
            rowsContainer.innerHTML = '';
            if (!student || !student.current_book_id || !startDay || !endDay || endDay < startDay) return;

            const roundNumber = parseInt(document.getElementById('f-round').value) || 1;
            const book = allData.books.find(b => b.id === student.current_book_id);
            let words = [];
            try {
                const res = await fetch(`/words?book_id=${student.current_book_id}&start_day=${startDay}&end_day=${endDay}&round_number=${roundNumber}`);
                words = await res.json();
            } catch (err) {
                console.error(err);
            }
            if (!words.length) {
                hint.textContent = '이 범위에 등록된 단어가 없어요 (엑셀 업로드가 필요할 수 있어요). 총 문항수를 Day별로 직접 입력해주세요.';
            } else if (book && book.pass_cutoff === null) {
                hint.textContent = `"${book.name}"에 통과 기준이 없어요 (엑셀 업로드로 자동 인식돼요). Day별로 합격 여부를 직접 체크해주세요.`;
            } else {
                hint.textContent = '';
            }

            const countByDay = {};
            words.forEach(w => { countByDay[w.day_number] = (countByDay[w.day_number] || 0) + 1; });

            let rowsHtml = '<table style="border-collapse:collapse; margin-top:6px;">' +
                '<tr><th style="text-align:left; font-size:12px; padding:2px 8px;">Day</th>' +
                '<th style="text-align:left; font-size:12px; padding:2px 8px;">총 문항수</th>' +
                '<th style="text-align:left; font-size:12px; padding:2px 8px;">틀린 개수</th>' +
                '<th style="text-align:left; font-size:12px; padding:2px 8px;">합격</th></tr>';
            for (let day = startDay; day <= endDay; day++) {
                const total = countByDay[day] || '';
                rowsHtml += `<tr data-day="${day}">` +
                    `<td style="padding:2px 8px; font-size:13px;">Day ${day}</td>` +
                    `<td style="padding:2px 8px;"><input type="number" class="fd-total" min="1" value="${total}" style="width:70px;" required></td>` +
                    `<td style="padding:2px 8px;"><input type="number" class="fd-wrong" min="0" value="0" style="width:70px;" required></td>` +
                    `<td style="padding:2px 8px;"><input type="checkbox" class="fd-passed"></td>` +
                    `</tr>`;
            }
            rowsHtml += '</table>';
            rowsContainer.innerHTML = rowsHtml;

            rowsContainer.querySelectorAll('.fd-wrong').forEach(input => {
                input.addEventListener('input', () => examRecalcRowPassed(input.closest('tr')));
                examRecalcRowPassed(input.closest('tr'));
            });
        }

        function examRecalcRowPassed(row) {
            const studentId = parseInt(document.getElementById('f-student').value);
            const student = allData.students.find(s => s.id === studentId);
            if (!student || !student.current_book_id) return;
            const book = allData.books.find(b => b.id === student.current_book_id);
            const wrong = parseInt(row.querySelector('.fd-wrong').value);
            if (book && book.pass_cutoff !== null && book.pass_cutoff !== undefined && !isNaN(wrong)) {
                row.querySelector('.fd-passed').checked = wrong <= book.pass_cutoff;
            }
        }

        document.getElementById('f-student').addEventListener('change', examAutoFillDay);
        document.getElementById('f-start-day').addEventListener('input', examBuildDayRows);
        document.getElementById('f-end-day').addEventListener('input', examBuildDayRows);
        document.getElementById('f-round').addEventListener('input', examBuildDayRows);

        document.getElementById('exam-form').addEventListener('submit', async (e) => {
            e.preventDefault();
            const studentId = parseInt(document.getElementById('f-student').value);
            const student = allData.students.find(s => s.id === studentId);
            if (!student || !student.current_book_id) {
                document.getElementById('f-status').textContent = '이 학생은 배정된 단어책이 없어요.';
                return;
            }
            const rows = document.querySelectorAll('#f-day-rows tr[data-day]');
            if (!rows.length) {
                document.getElementById('f-status').textContent = 'Day 범위를 먼저 입력해주세요.';
                return;
            }
            const roundNumber = parseInt(document.getElementById('f-round').value);
            const examDate = document.getElementById('f-date').value;
            const results = [];
            for (const row of rows) {
                const day = parseInt(row.dataset.day);
                const total = parseInt(row.querySelector('.fd-total').value);
                const wrong = parseInt(row.querySelector('.fd-wrong').value) || 0;
                const passed = row.querySelector('.fd-passed').checked;
                const score = Math.max(0, total - wrong);
                const body = {
                    student_id: studentId,
                    book_id: student.current_book_id,
                    round_number: roundNumber,
                    start_day: day,
                    end_day: day,
                    exam_date: examDate,
                    score: score,
                    total_questions: total,
                    passed: passed,
                };
                try {
                    const res = await fetch('/exam-records', {
                        method: 'POST',
                        headers: { 'Content-Type': 'application/json' },
                        body: JSON.stringify(body),
                    });
                    if (res.ok) {
                        results.push(`Day${day}:${passed ? '합격' : '재시험'}`);
                    } else {
                        const errText = await res.text();
                        let detail = errText;
                        try { detail = JSON.parse(errText).detail || errText; } catch (e) {}
                        results.push(`Day${day}:실패(${detail})`);
                    }
                } catch (err) {
                    results.push(`Day${day}:서버연결실패`);
                    console.error(err);
                }
            }
            document.getElementById('f-status').textContent = results.join(', ');
            document.getElementById('f-start-day').value = '';
            document.getElementById('f-end-day').value = '';
            document.getElementById('f-day-rows').innerHTML = '';
            document.getElementById('f-hint').textContent = '';
            loadData();
        });

        document.querySelectorAll('.tab-btn').forEach(btn => {
            btn.addEventListener('click', () => {
                document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
                document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
                btn.classList.add('active');
                document.getElementById('tab-' + btn.dataset.tab).classList.add('active');
                if (btn.dataset.tab === 'history') renderCalendar();
            });
        });

        const calToday = new Date();
        let calYear = calToday.getFullYear();
        let calMonth = calToday.getMonth();

        function examRecordsByDate() {
            const map = {};
            (allData.examRecords || []).forEach(r => {
                if (!map[r.exam_date]) map[r.exam_date] = [];
                map[r.exam_date].push(r);
            });
            return map;
        }

        function renderCalendar() {
            const grid = document.getElementById('calendar-grid');
            document.getElementById('cal-label').textContent = `${calYear}년 ${calMonth + 1}월`;

            const byDate = examRecordsByDate();
            const firstWeekday = new Date(calYear, calMonth, 1).getDay();
            const daysInMonth = new Date(calYear, calMonth + 1, 0).getDate();

            let html = ['일', '월', '화', '수', '목', '금', '토']
                .map(d => `<div class="cal-weekday">${d}</div>`).join('');
            for (let i = 0; i < firstWeekday; i++) {
                html += '<div class="cal-cell empty"></div>';
            }
            for (let d = 1; d <= daysInMonth; d++) {
                const dateStr = `${calYear}-${String(calMonth + 1).padStart(2, '0')}-${String(d).padStart(2, '0')}`;
                const records = byDate[dateStr] || [];
                const hasData = records.length > 0;
                html += `<div class="cal-cell ${hasData ? 'has-data' : ''}" data-date="${dateStr}">` +
                    `<div class="cal-date">${d}</div>` +
                    (hasData ? `<div class="cal-count">${records.length}건</div>` : '') +
                    `</div>`;
            }
            grid.innerHTML = html;

            grid.querySelectorAll('.cal-cell.has-data').forEach(cell => {
                cell.addEventListener('click', () => showDayDetail(cell.dataset.date));
            });
        }

        function showDayDetail(dateStr) {
            document.querySelectorAll('.cal-cell').forEach(c => c.classList.remove('selected'));
            const cell = document.querySelector(`.cal-cell[data-date="${dateStr}"]`);
            if (cell) cell.classList.add('selected');

            const byDate = examRecordsByDate();
            const records = byDate[dateStr] || [];
            const studentMap = Object.fromEntries(allData.students.map(s => [s.id, s]));
            const bookMap = Object.fromEntries(allData.books.map(b => [b.id, b.name]));
            const detail = document.getElementById('day-detail');

            if (!records.length) {
                detail.innerHTML = `<p class="empty">${dateStr}에 등록된 시험 기록이 없습니다.</p>`;
                return;
            }
            const passCount = records.filter(r => r.passed).length;
            const rowsHtml = records.map(r => {
                const student = studentMap[r.student_id];
                return `<tr>
                    <td>${student ? student.name : '-'}</td>
                    <td>${bookMap[r.book_id] || '-'}</td>
                    <td>${r.round_number}회독</td>
                    <td>Day ${r.start_day}~${r.end_day}</td>
                    <td>${r.score ?? '-'}/${r.total_questions ?? '-'}</td>
                    <td>${r.passed ? '합격' : '재시험'}</td>
                    <td><a class="btn" href="/exam-records/${r.id}/retest-sheet" target="_blank">시험지 보기</a></td>
                </tr>`;
            }).join('');
            detail.innerHTML = `
                <h2>${dateStr} (${records.length}건 · 합격 ${passCount} · 재시험 ${records.length - passCount})</h2>
                <table>
                    <tr><th>학생</th><th>단어책</th><th>회독</th><th>범위</th><th>점수</th><th>결과</th><th>시험지</th></tr>
                    ${rowsHtml}
                </table>
            `;
        }

        document.getElementById('cal-prev').addEventListener('click', () => {
            calMonth--;
            if (calMonth < 0) { calMonth = 11; calYear--; }
            renderCalendar();
        });
        document.getElementById('cal-next').addEventListener('click', () => {
            calMonth++;
            if (calMonth > 11) { calMonth = 0; calYear++; }
            renderCalendar();
        });

        loadData();
    </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html, headers={"Cache-Control": "no-store"})

# ---------------------------------------------------------------
# Teacher
# ---------------------------------------------------------------
@app.post("/teachers", response_model=schemas.TeacherOut)
def create_teacher(payload: schemas.TeacherCreate, db: Session = Depends(get_db)):
    existing = db.query(models.Teacher).filter(models.Teacher.username == payload.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="이미 사용 중인 아이디입니다")
    teacher = models.Teacher(
        name=payload.name, username=payload.username, password_hash=payload.password
    )
    db.add(teacher)
    db.commit()
    db.refresh(teacher)
    return teacher


@app.get("/teachers", response_model=List[schemas.TeacherOut])
def list_teachers(db: Session = Depends(get_db)):
    return db.query(models.Teacher).all()


# ---------------------------------------------------------------
# Book
# ---------------------------------------------------------------
@app.post("/books", response_model=schemas.BookOut)
def create_book(payload: schemas.BookCreate, db: Session = Depends(get_db)):
    book = models.Book(**payload.model_dump())
    db.add(book)
    db.commit()
    db.refresh(book)
    return book


@app.get("/books", response_model=List[schemas.BookOut])
def list_books(db: Session = Depends(get_db)):
    return db.query(models.Book).all()


@app.post("/books/bulk", response_model=List[schemas.BookOut])
def create_books_bulk(payload: schemas.BookBulkCreate, db: Session = Depends(get_db)):
    existing_names = {b.name for b in db.query(models.Book.name).all()}
    new_books = []
    for name in payload.names:
        if name in existing_names:
            continue
        book = models.Book(name=name, publisher=payload.publisher)
        db.add(book)
        new_books.append(book)
        existing_names.add(name)

    db.commit()
    for book in new_books:
        db.refresh(book)
    return new_books


# ---------------------------------------------------------------
# Student
# ---------------------------------------------------------------
@app.post("/students", response_model=schemas.StudentOut)
def create_student(payload: schemas.StudentCreate, db: Session = Depends(get_db)):
    student = models.Student(**payload.model_dump())
    db.add(student)
    db.commit()
    db.refresh(student)
    return student


@app.post("/students/bulk", response_model=List[schemas.StudentOut])
def create_students_bulk(payload: schemas.StudentBulkCreate, db: Session = Depends(get_db)):
    new_students = []
    for name in payload.names:
        student = models.Student(name=name)
        db.add(student)
        new_students.append(student)

    db.commit()
    for student in new_students:
        db.refresh(student)
    return new_students


@app.patch("/students/{student_id}", response_model=schemas.StudentOut)
def update_student(
    student_id: int, payload: schemas.StudentUpdate, db: Session = Depends(get_db)
):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    update_data = payload.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(student, field, value)

    db.commit()
    db.refresh(student)
    return student


@app.get("/students", response_model=List[schemas.StudentOut])
def list_students(
    teacher_id: Optional[int] = None,
    include_inactive: bool = False,
    db: Session = Depends(get_db),
):
    query = db.query(models.Student)
    if teacher_id is not None:
        query = query.filter(models.Student.teacher_id == teacher_id)
    if not include_inactive:
        query = query.filter(models.Student.is_active == True)
    return query.all()


@app.get("/students/{student_id}", response_model=schemas.StudentOut)
def get_student(student_id: int, db: Session = Depends(get_db)):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")
    return student


@app.delete("/students/{student_id}")
def delete_student(student_id: int, db: Session = Depends(get_db)):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")
    db.delete(student)
    db.commit()
    return {"deleted": True, "id": student_id}


# ---------------------------------------------------------------
# Schedule (자동 스케줄 - 오늘 시험 범위)
# ---------------------------------------------------------------
def get_absent_dates(db, student_id, up_to_date):
    rows = (
        db.query(models.Attendance.date)
        .filter(models.Attendance.student_id == student_id, models.Attendance.date <= up_to_date)
        .all()
    )
    return {r.date for r in rows}


@app.get("/students/{student_id}/today-session")
def student_today_session(
    student_id: int, on_date: Optional[str] = None, db: Session = Depends(get_db)
):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    book = db.query(models.Book).filter(models.Book.id == student.current_book_id).first() if student.current_book_id else None
    target_date = date.fromisoformat(on_date) if on_date else date.today()
    absent_dates = get_absent_dates(db, student_id, target_date)
    session = compute_today_session(student, target_date, total_days=book.total_days if book else None, absent_dates=absent_dates)
    return {"student_id": student_id, "date": target_date.isoformat(), "session": session}


@app.get("/classes/{class_name}/today-sessions")
def class_today_sessions(
    class_name: str, on_date: Optional[str] = None, db: Session = Depends(get_db)
):
    target_date = date.fromisoformat(on_date) if on_date else date.today()
    students = (
        db.query(models.Student)
        .filter(models.Student.class_name == class_name, models.Student.is_active == True)
        .all()
    )

    result = []
    for student in students:
        book = db.query(models.Book).filter(models.Book.id == student.current_book_id).first() if student.current_book_id else None
        absent_dates = get_absent_dates(db, student.id, target_date)
        session = compute_today_session(student, target_date, total_days=book.total_days if book else None, absent_dates=absent_dates)
        if session:
            result.append(
                {
                    "student_id": student.id,
                    "name": student.name,
                    "type": session["type"],
                    "start_day": session["start_day"],
                    "end_day": session["end_day"],
                    "round": session["round"],
                }
            )
    return {"class_name": class_name, "date": target_date.isoformat(), "sessions": result}


@app.get("/classes/{class_name}/attendance")
def class_attendance(
    class_name: str, on_date: Optional[str] = None, db: Session = Depends(get_db)
):
    target_date = date.fromisoformat(on_date) if on_date else date.today()
    students = (
        db.query(models.Student)
        .filter(models.Student.class_name == class_name, models.Student.is_active == True)
        .all()
    )
    student_ids = [s.id for s in students]
    absent_rows = (
        db.query(models.Attendance.student_id)
        .filter(models.Attendance.student_id.in_(student_ids), models.Attendance.date == target_date)
        .all()
    )
    absent_ids = {r.student_id for r in absent_rows}
    return {
        "class_name": class_name,
        "date": target_date.isoformat(),
        "attendance": {s.id: (s.id not in absent_ids) for s in students},
    }


@app.post("/attendance")
def set_attendance(payload: schemas.AttendanceSet, db: Session = Depends(get_db)):
    existing = (
        db.query(models.Attendance)
        .filter(models.Attendance.student_id == payload.student_id, models.Attendance.date == payload.date)
        .first()
    )
    if payload.present:
        if existing:
            db.delete(existing)
            db.commit()
    else:
        if not existing:
            db.add(models.Attendance(student_id=payload.student_id, date=payload.date))
            db.commit()
    return {"student_id": payload.student_id, "date": payload.date.isoformat(), "present": payload.present}


# ---------------------------------------------------------------
# FinalReviewFile (최종시험지 원본 파일)
# ---------------------------------------------------------------
@app.post("/final-review-files", response_model=schemas.FinalReviewFileOut)
async def upload_final_review_file(
    book_id: int = Form(...),
    start_day: int = Form(...),
    end_day: int = Form(...),
    version_label: str = Form("A"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    book = db.query(models.Book).filter(models.Book.id == book_id).first()
    if not book:
        raise HTTPException(status_code=404, detail="단어책을 찾을 수 없습니다")

    ext = os.path.splitext(file.filename)[1]
    stored_name = f"{uuid.uuid4().hex}{ext}"
    stored_path = os.path.join(FINAL_REVIEW_DIR, stored_name)
    content = await file.read()
    with open(stored_path, "wb") as f:
        f.write(content)

    record = models.FinalReviewFile(
        book_id=book_id,
        start_day=start_day,
        end_day=end_day,
        version_label=version_label,
        file_path=stored_name,
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@app.get("/final-review-files", response_model=List[schemas.FinalReviewFileOut])
def list_final_review_files(
    book_id: Optional[int] = None,
    start_day: Optional[int] = None,
    end_day: Optional[int] = None,
    db: Session = Depends(get_db),
):
    query = db.query(models.FinalReviewFile)
    if book_id is not None:
        query = query.filter(models.FinalReviewFile.book_id == book_id)
    if start_day is not None:
        query = query.filter(models.FinalReviewFile.start_day == start_day)
    if end_day is not None:
        query = query.filter(models.FinalReviewFile.end_day == end_day)
    return query.all()


@app.get("/final-review-files/{file_id}/download")
def download_final_review_file(file_id: int, db: Session = Depends(get_db)):
    record = db.query(models.FinalReviewFile).filter(models.FinalReviewFile.id == file_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다")
    full_path = os.path.join(FINAL_REVIEW_DIR, record.file_path)
    if not os.path.exists(full_path):
        raise HTTPException(status_code=404, detail="저장된 파일이 없습니다")
    book = db.query(models.Book).filter(models.Book.id == record.book_id).first()
    book_name = book.name if book else "book"
    original_ext = os.path.splitext(record.file_path)[1]
    download_name = f"{book_name} {record.start_day}~{record.end_day} {record.version_label}{original_ext}"
    return FileResponse(full_path, filename=download_name)


@app.delete("/final-review-files/{file_id}")
def delete_final_review_file(file_id: int, db: Session = Depends(get_db)):
    record = db.query(models.FinalReviewFile).filter(models.FinalReviewFile.id == file_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다")
    full_path = os.path.join(FINAL_REVIEW_DIR, record.file_path)
    if os.path.exists(full_path):
        os.remove(full_path)
    db.delete(record)
    db.commit()
    return {"deleted": True, "id": file_id}


# ---------------------------------------------------------------
# Word
# ---------------------------------------------------------------
@app.post("/words/upload-excel")
async def upload_words_excel(
    book_id: int = Form(...),
    sheet_name: Optional[str] = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    book = db.query(models.Book).filter(models.Book.id == book_id).first()
    if not book:
        raise HTTPException(status_code=404, detail="단어책을 찾을 수 없습니다")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    pairs = []
    seen_pairs = set()
    left_day = None
    right_day = None
    left_round = 1
    right_round = 1
    max_day_seen = 0
    cutoff_value = None

    day_header_re = re.compile(r"day\s*(\d+)(?:\s*\(\s*(\d+)\s*회독\s*\))?", re.IGNORECASE)

    for row in ws.iter_rows(values_only=True):
        row = (row + (None,) * 6)[:6]
        c0, c1, c2, c3, c4, c5 = row

        if c0 == "No.":
            m1 = day_header_re.search(str(c1)) if c1 else None
            m2 = day_header_re.search(str(c4)) if c4 else None
            left_day = int(m1.group(1)) if m1 else None
            left_round = int(m1.group(2)) if (m1 and m1.group(2)) else 1
            right_day = int(m2.group(1)) if m2 else None
            right_round = int(m2.group(2)) if (m2 and m2.group(2)) else 1
            if left_day:
                max_day_seen = max(max_day_seen, left_day)
            if right_day:
                max_day_seen = max(max_day_seen, right_day)

            if cutoff_value is None:
                for cutoff_cell in (c2, c5):
                    if cutoff_cell:
                        mc = re.search(r"통과\s*(?:기준)?\s*:\s*-?(\d+)", str(cutoff_cell))
                        if mc:
                            cutoff_value = int(mc.group(1))
                            break
            continue

        left_content = c1 if c1 else c2
        if left_content and left_day:
            key = (left_day, left_round, str(left_content).strip())
            if key not in seen_pairs:
                seen_pairs.add(key)
                pairs.append(key)
        right_content = c4 if c4 else c5
        if right_content and right_day:
            key = (right_day, right_round, str(right_content).strip())
            if key not in seen_pairs:
                seen_pairs.add(key)
                pairs.append(key)

    if not pairs:
        raise HTTPException(
            status_code=422,
            detail="엑셀에서 단어를 찾지 못했습니다. 시트/양식을 확인해주세요.",
        )

    if max_day_seen:
        book.total_days = max_day_seen
    if cutoff_value is not None:
        book.pass_cutoff = cutoff_value

    existing = {
        (w.day_number, w.round_number, w.word)
        for w in db.query(models.Word).filter(models.Word.book_id == book_id).all()
    }

    new_words = []
    for day_number, round_number, word in pairs:
        if (day_number, round_number, word) in existing:
            continue
        w = models.Word(book_id=book_id, day_number=day_number, round_number=round_number, word=word)
        db.add(w)
        new_words.append(w)

    db.commit()

    days = sorted(set(d for d, _, _ in pairs))
    rounds = sorted(set(r for _, r, _ in pairs))
    return {
        "book_id": book_id,
        "parsed_pairs": len(pairs),
        "newly_inserted": len(new_words),
        "skipped_duplicates": len(pairs) - len(new_words),
        "day_range": {"min": min(days), "max": max(days)} if days else None,
        "rounds_detected": rounds,
        "total_days_detected": max_day_seen or None,
        "pass_cutoff_detected": cutoff_value,
    }

@app.post("/words", response_model=schemas.WordOut)
def create_word(payload: schemas.WordCreate, db: Session = Depends(get_db)):
    word = models.Word(**payload.model_dump())
    db.add(word)
    db.commit()
    db.refresh(word)
    return word


@app.get("/words", response_model=List[schemas.WordOut])
def list_words(
    book_id: int,
    start_day: Optional[int] = None,
    end_day: Optional[int] = None,
    round_number: int = 1,
    db: Session = Depends(get_db),
):
    query = db.query(models.Word).filter(
        models.Word.book_id == book_id, models.Word.round_number == round_number
    )
    if start_day is not None:
        query = query.filter(models.Word.day_number >= start_day)
    if end_day is not None:
        query = query.filter(models.Word.day_number <= end_day)
    return query.order_by(models.Word.day_number, models.Word.id).all()


# ---------------------------------------------------------------
# ExamRecord
# ---------------------------------------------------------------
@app.post("/exam-records", response_model=schemas.ExamRecordOut)
def create_exam_record(payload: schemas.ExamRecordCreate, db: Session = Depends(get_db)):
    record = models.ExamRecord(**payload.model_dump())
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@app.get("/exam-records", response_model=List[schemas.ExamRecordOut])
def list_exam_records(
    student_id: Optional[int] = None, db: Session = Depends(get_db)
):
    query = db.query(models.ExamRecord)
    if student_id is not None:
        query = query.filter(models.ExamRecord.student_id == student_id)
    return query.order_by(models.ExamRecord.exam_date.desc()).all()


@app.get("/exam-records/retest-candidates")
def retest_candidates(
    due_only: bool = False,
    on_date: Optional[str] = None,
    db: Session = Depends(get_db),
):
    failed = db.query(models.ExamRecord).filter(models.ExamRecord.passed == False).all()
    retaken_ids = {
        r.original_exam_id
        for r in db.query(models.ExamRecord).filter(
            models.ExamRecord.original_exam_id.isnot(None)
        )
    }
    candidates = [r for r in failed if r.id not in retaken_ids]

    target_date = date.fromisoformat(on_date) if on_date else date.today()

    result = []
    for r in candidates:
        student = db.query(models.Student).filter(models.Student.id == r.student_id).first()
        due = next_attend_date(r.exam_date, student.attend_days if student else None)
        if due_only and due != target_date:
            continue
        item = schemas.ExamRecordOut.model_validate(r).model_dump(mode="json")
        item["due_date"] = due.isoformat() if due else None
        result.append(item)
    return result


@app.get("/exam-records/{exam_record_id}/retest-sheet", response_class=HTMLResponse)
def retest_sheet(exam_record_id: int, db: Session = Depends(get_db)):
    record = db.query(models.ExamRecord).filter(models.ExamRecord.id == exam_record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="시험 기록을 찾을 수 없습니다")

    student = db.query(models.Student).filter(models.Student.id == record.student_id).first()
    book = db.query(models.Book).filter(models.Book.id == record.book_id).first()

    words = get_ordered_words(db, record.book_id, record.start_day, record.end_day, record.round_number)
    if not words:
        raise HTTPException(status_code=422, detail="이 범위에 해당하는 단어가 DB에 없습니다. 단어 업로드를 먼저 확인해주세요.")

    html = f"""
    <html>
    <head>
    <meta charset="utf-8">
    <title>재시험지</title>
    <style>{RETEST_SHEET_STYLE}</style>
    </head>
    <body>
        <button onclick="window.print()">인쇄하기</button>
        <h1>재시험지 ({record.round_number}회독)</h1>
        <div class="meta">
            학생: {student.name if student else '-'} &nbsp;|&nbsp;
            단어책: {book.name if book else '-'} &nbsp;|&nbsp;
            범위: Day {record.start_day}~{record.end_day}
        </div>
        {render_word_table(words)}
    </body>
    </html>
    """
    return HTMLResponse(content=html)


@app.get("/students/{student_id}/retest-sheet", response_class=HTMLResponse)
def student_combined_retest_sheet(student_id: int, db: Session = Depends(get_db)):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    failed = (
        db.query(models.ExamRecord)
        .filter(
            models.ExamRecord.student_id == student_id,
            models.ExamRecord.passed == False,
        )
        .all()
    )
    retaken_ids = {
        r.original_exam_id
        for r in db.query(models.ExamRecord).filter(
            models.ExamRecord.original_exam_id.isnot(None)
        )
    }
    candidates = [r for r in failed if r.id not in retaken_ids]

    if not candidates:
        return HTMLResponse(content=f"""
        <html><head><meta charset="utf-8"></head>
        <body style="font-family: sans-serif; padding: 24px;">
            <h2>{student.name} - 재시험 대상이 없습니다</h2>
            <p style="color: #888;">현재 재시험이 필요한 시험 기록이 없어요.</p>
        </body></html>
        """)

    candidates.sort(key=lambda r: (r.round_number, r.start_day))

    sections = []
    for record in candidates:
        book = db.query(models.Book).filter(models.Book.id == record.book_id).first()
        words = get_ordered_words(db, record.book_id, record.start_day, record.end_day, record.round_number)
        if not words:
            continue
        title = (
            f"{record.round_number}회독 Day {record.start_day}~{record.end_day} 재시험 "
            f"<span class='book'>({book.name if book else '-'})</span>"
        )
        sections.append((title, words))

    html = f"""
    <html>
    <head>
    <meta charset="utf-8">
    <title>{student.name} 통합 재시험지</title>
    <style>{RETEST_SHEET_STYLE}</style>
    </head>
    <body>
        <button onclick="window.print()">인쇄하기</button>
        <h1>{student.name} 재시험지 (재시험 {len(candidates)}건)</h1>
        {render_day_sections(sections)}
    </body>
    </html>
    """
    return HTMLResponse(content=html)


@app.get("/classes/{class_name}/retest-print-queue", response_class=HTMLResponse)
def class_retest_print_queue(
    class_name: str, on_date: Optional[str] = None, db: Session = Depends(get_db)
):
    target_date = date.fromisoformat(on_date) if on_date else date.today()

    students = (
        db.query(models.Student).filter(models.Student.class_name == class_name).all()
    )
    if not students:
        raise HTTPException(status_code=404, detail="해당 반의 학생을 찾을 수 없습니다")

    failed = db.query(models.ExamRecord).filter(models.ExamRecord.passed == False).all()
    retaken_ids = {
        r.original_exam_id
        for r in db.query(models.ExamRecord).filter(
            models.ExamRecord.original_exam_id.isnot(None)
        )
    }
    all_candidates = [r for r in failed if r.id not in retaken_ids]

    pages_html = ""
    included_count = 0

    for student in students:
        student_candidates = [
            r
            for r in all_candidates
            if r.student_id == student.id
            and next_attend_date(r.exam_date, student.attend_days) == target_date
        ]
        if not student_candidates:
            continue

        student_candidates.sort(key=lambda r: (r.round_number, r.start_day))
        included_count += 1

        sections = []
        for record in student_candidates:
            book = db.query(models.Book).filter(models.Book.id == record.book_id).first()
            words = get_ordered_words(db, record.book_id, record.start_day, record.end_day, record.round_number)
            if not words:
                continue
            title = (
                f"{record.round_number}회독 Day {record.start_day}~{record.end_day} 재시험 "
                f"<span class='book'>({book.name if book else '-'})</span>"
            )
            sections.append((title, words))

        pages_html += f"""
        <div class="page">
            <h1>{student.name} <span class="class">({class_name})</span></h1>
            {render_day_sections(sections)}
        </div>
        """

    if included_count == 0:
        return HTMLResponse(content=f"""
        <html><head><meta charset="utf-8"></head>
        <body style="font-family: sans-serif; padding: 24px;">
            <h2>{class_name} - {target_date.isoformat()}에 인쇄할 재시험지가 없습니다</h2>
        </body></html>
        """)

    html = f"""
    <html>
    <head>
    <meta charset="utf-8">
    <title>{class_name} 재시험지 모음 ({target_date.isoformat()})</title>
    <style>{RETEST_SHEET_STYLE}</style>
    </head>
    <body>
        <button onclick="window.print()">전체 인쇄하기</button>
        <p style="color:#666;">{class_name} 반 · {target_date.isoformat()} 인쇄 대상 · 학생 {included_count}명</p>
        {pages_html}
    </body>
    </html>
    """
    return HTMLResponse(content=html)